from openpyxl import load_workbook

class ExcelReader:
    def __init__(self, excel_file_path):
        self.excel_file_path = excel_file_path

    def read_testcase_data(self, row_number, column_name, sheet_name="Sheet1"):
        workbook = load_workbook(filename=self.excel_file_path)
        sheet = workbook[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        col_index = headers.index(column_name) + 1
        cell_value = sheet.cell(row=row_number, column=col_index).value

        return str(cell_value) if cell_value is not None else None
